"""
PDF → Excel 結案表填充器
功能:上傳銀河專案 PDF,自動填入結案表模板對應位置
"""
import base64

# ==================== 內建模板資料 ====================
"""內建模板資料(base64 編碼)"""

CLOSURE_TEMPLATE_B64 = """UEsEMwYAAAAAFwAXACIGAABRegEAAAA=""" 
MONTHLY_TEMPLATE_B64 = """UEsFBYAAAAAFwAXACIGAABRegEAAAA=""" 

# ==================== 內建模板載入函式 ====================

def get_builtin_template(report_type):
    """取得內建模板的 bytes 資料"""
    if report_type == "月報表":
        return base64.b64decode(MONTHLY_TEMPLATE_B64)
    elif report_type == "結案表":
        return base64.b64decode(CLOSURE_TEMPLATE_B64)
    raise ValueError(f"未知的報表類型: {report_type}")


import streamlit as st
import pdfplumber
import re
import unicodedata
import subprocess
import tempfile
import os
import shutil
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime

st.set_page_config(page_title="PDF → 報表填充器", page_icon="📊", layout="wide")

# ==================== PDF 抽取邏輯 ====================

def normalize(s):
    """去空白、Unicode 正規化(把 ⽂→文 等異體字統一)"""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize('NFKC', s)
    return re.sub(r'\s+', '', s)


def is_main_data_row(row):
    """判斷是否為主表資料列"""
    if not row or len(row) < 13:
        return False
    站版 = normalize(row[0])
    if not 站版 or 站版 == "合計" or "站版" in 站版:
        return False
    try:
        int(row[4])
        return True
    except (ValueError, TypeError):
        return False


# 主軸縮寫對照(PDF 中主軸欄位常被切成單字)
主軸縮寫對照 = {
    "自": "自品置入",
    "需": "需求置入",
    "圖": "圖書/閱讀",
    "會": "會員/品牌",
    "百": "百貨",
    "競": "競品置入",
    "全": "全域置入",
    "社": "社團置入",
    "消": "消費者需求置入",
    "情": "情境話題",
    "產": "產品話題",
    "話": "話題",
}

# 站版前綴關鍵字 — 用來辨識一行是否以「站版」開頭
站版前綴 = ["Threads", "PTT", "Facebook", "Dcard", "Mobile01", "eyny", "Plurk",
             "BabyHome", "FG", "伊莉", "狄卡", "批踢踢", "FashionGuide"]

# 行尾「主軸 + 日期前半 + 9 個數字」的通用特徵
ROW_PATTERN = re.compile(
    r'([^\s\d][^\s]*)\s+'                                          # 主軸縮寫(非數字、非空白開頭)
    r'\d{4}-\s+'                                                   # 日期 yyyy-
    r'(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+'                            # 發文 回應 聲量 正向
    r'(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)'                       # 正面 負面 議題 產品 複合
    r'\s*$', re.MULTILINE
)


def get_main_category_from_pdf(pdf_path):
    """從 PDF「操作主軸」區塊抽出實際主軸全名清單"""
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += unicodedata.normalize('NFKC', page.extract_text() or "") + "\n"
            
    extracted = []
    
    start_idx = full_text.find("操作主軸")
    if start_idx < 0:
        return extracted
        
    end_markers = ["專案聲量分佈", "網友關注度", "網友回應概況"]
    end_idx = len(full_text)
    for marker in end_markers:
        idx = full_text.find(marker, start_idx + 10)
        if 0 < idx < end_idx:
            end_idx = idx
            
    section = full_text[start_idx:end_idx]
    
    for line in section.split('\n'):
        line = line.strip()
        if not line:
            continue
        if line in ("操作主軸", "類型 累計", "類型", "累計"):
            continue
        m = re.match(r'^(.+?)\s+(\d+)\s*$', line)
        if m:
            name = m.group(1).strip()
            if name == "合計":
                continue
            if ":" in name or "%" in name:
                continue
            if re.search(r'\d+\.\d+', line):
                continue
            extracted.append(name)
            
    return extracted


def resolve_main_category(prefix, full_names):
    """把截斷的主軸縮寫對應到完整名稱"""
    if not prefix:
        return prefix
    for name in full_names:
        if name == prefix:
            return name
    for name in full_names:
        if name.startswith(prefix):
            return name
    for name in full_names:
        if prefix.startswith(name) and len(name) >= 2:
            return name
    return prefix


def extract_main_table_via_text(pdf_path):
    """用「行尾 9 個數字」特徵從純文字中抓主表資料"""
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            full_text += unicodedata.normalize('NFKC', page_text) + "\n"
            
    full_names = get_main_category_from_pdf(pdf_path)
    
    rows = []
    for m in ROW_PATTERN.finditer(full_text):
        主軸縮寫 = m.group(1)
        nums = [int(m.group(i)) for i in range(2, 11)]
        
        line_start = full_text.rfind('\n', 0, m.start()) + 1
        line_part = full_text[line_start:m.start()].strip()
        
        if len(line_part) < 3:
            continue
            
        主軸全名 = resolve_main_category(主軸縮寫, full_names)
        if 主軸全名 == 主軸縮寫 and 主軸縮寫 in 主軸縮寫對照:
            主軸全名 = 主軸縮寫對照[主軸縮寫]
            
        rows.append({
            "站版": line_part,
            "標題": "",
            "主軸": 主軸全名,
            "專案發文量": nums[0],
            "網友回應量": nums[1],
            "討論聲量總數": nums[2],
            "正向聲量總數": nums[3],
            "正面討論": nums[4],
            "負面討論": nums[5],
            "議題討論": nums[6],
            "產品討論": nums[7],
            "複合討論": nums[8],
        })
    return rows


def extract_main_table(pdf_path):
    """抽取主表"""
    rows = extract_main_table_via_text(pdf_path)
    if rows:
        return rows
        
    fallback_rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table:
                    continue
                for row in table:
                    if is_main_data_row(row):
                        fallback_rows.append({
                            "站版": normalize(row[0]),
                            "標題": (row[1] or "").strip().replace('\n', ' '),
                            "主軸": normalize(row[2]),
                            "專案發文量": int(row[4]),
                            "網友回應量": int(row[5]),
                            "討論聲量總數": int(row[6]),
                            "正向聲量總數": int(row[7]),
                            "正面討論": int(row[8]),
                            "負面討論": int(row[9]),
                            "議題討論": int(row[10]),
                            "產品討論": int(row[11]),
                            "複合討論": int(row[12]),
                        })
    return fallback_rows


def extract_post_types(pdf_path):
    """抽取「文案類型發文篇數」表格"""
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += unicodedata.normalize('NFKC', page.extract_text() or "") + "\n"
            
    types = {}
    start_idx = full_text.find("文案類型發文篇數")
    if start_idx < 0:
        start_idx = full_text.find("文案類型")
    if start_idx < 0:
        return types
        
    end_markers = ["專案執行進度摘要", "操作主軸", "專案聲量分佈", "網友關注度", "網友回應概況"]
    end_idx = len(full_text)
    for marker in end_markers:
        idx = full_text.find(marker, start_idx + 10)
        if 0 < idx < end_idx:
            end_idx = idx
            
    section = full_text[start_idx:end_idx]
    
    for line in section.split('\n'):
        line = line.strip()
        if not line:
            continue
        if line in ("文案類型 累計", "文案類型", "累計"):
            continue
        m = re.match(r'^(.+?)\s+(\d+)\s*$', line)
        if m:
            name = m.group(1).strip()
            count = int(m.group(2))
            if name == "合計":
                continue
            if ":" in name or "%" in name:
                continue
            if re.search(r'\d+\.\d+', line):
                continue
            types[name] = count
            
    return types


def extract_posts_with_content(pdf_path):
    """從「專案發文總覽 PDF」智慧切分並還原斷行與碎裂的標題"""
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for p in pdf.pages:
            full_text += unicodedata.normalize('NFKC', p.extract_text() or '') + "\n"
            
    # 利用每篇開頭必有的「#數字 RE:」或「■ #數字 RE:」做大範圍的區塊切分
    post_start_pattern = r'(?:title\.標題\s+)?(?=(?:■\s*)?#\d+\s*RE\s*[:：\s])'
    post_matches = list(re.finditer(post_start_pattern, full_text, re.IGNORECASE))
    
    posts = []
    for i, m in enumerate(post_matches):
        start_pos = m.start()
        end_pos = post_matches[i + 1].start() if i + 1 < len(post_matches) else len(full_text)
        
        # 取得單一文章區塊文本
        block = full_text[start_pos:end_pos]
        
        # 鎖定標題範圍：從區塊開頭到第一個「cnt.內文」或「類型:」等關鍵字前的所有文字碎片
        title_part_match = re.search(r'^(.*?)(?=\n\s*(?:類型|網址紀錄|文案詳細內容|cnt\.內文|$))', block, re.DOTALL | re.IGNORECASE)
        
        if title_part_match:
            raw_title = title_part_match.group(1).strip()
            # 抹除可能包含的 "title.標題" 關鍵字
            raw_title = re.sub(r'^title\.標題\s*', '', raw_title, flags=re.IGNORECASE)
            # 將所有因排版錯位產生的換行與多重空白，完美合併還原為單一標準空格
            title = re.sub(r'\s+', ' ', raw_title)
        else:
            lines = [l.strip() for l in block.split('\n') if l.strip()]
            title = lines[0] if lines else "未知標題"
            
        # 擷取內文：從 cnt.內文 開始，一直到「截圖」或下個區塊結束
        content_match = re.search(r'cnt\.內文\s*(.*?)(?:\n\s*截圖|\n\s*Time\.發文時間|\Z)', block, re.DOTALL | re.IGNORECASE)
        content = content_match.group(1).strip() if content_match else ""
        
        if title or content:
            posts.append({"title": title, "content": content})
            
    return posts


def normalize_thread_title(title):
    """把標題去掉 #N RE: 前綴 + 過濾 PDF 圖示字符,得到主討論串名稱"""
    if not title:
        return ""
        
    cleaned = title
    cleaned = ''.join(ch for ch in cleaned if not (0xE000 <= ord(ch) <= 0xF8FF))
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    # 移除回文前綴（包含半形冒號、全形冒號與不分大小寫）
    cleaned = re.sub(r'^#\d*\s*RE\s*[:：]\s*', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'^RE\s*[:：]\s*', '', cleaned, flags=re.IGNORECASE)
    
    return cleaned.strip()


def search_posts(posts, keyword, search_in="content"):
    """依關鍵字搜尋文章"""
    if not keyword.strip():
        return []
        
    kw = keyword.strip()
    if '+' in kw:
        parts = [k.strip().lower() for k in kw.split('+') if k.strip()]
        mode = "AND"
    elif ',' in kw or '|' in kw:
        parts = [k.strip().lower() for k in re.split(r'[,|]', kw) if k.strip()]
        mode = "OR"
    else:
        parts = [kw.lower()]
        mode = "AND"
        
    def match(text):
        t = text.lower()
        return all(k in t for k in parts) if mode == "AND" else any(k in t for k in parts)
        
    if search_in == "title":
        from collections import OrderedDict
        groups = OrderedDict()
        for p in posts:
            key = normalize_thread_title(p["title"])
            if key not in groups:
                groups[key] = []
            groups[key].append(p)
            
        matched = []
        for thread_title, group_posts in groups.items():
            if match(thread_title):
                rep = group_posts[0]
                matched.append({
                    "title": thread_title,
                    "content": rep["content"],
                    "snippet_source": thread_title,
                    "thread_size": len(group_posts),
                })
        return matched
    else:
        matched = []
        for p in posts:
            if match(p["content"]):
                matched.append({
                    "title": p["title"],
                    "content": p["content"],
                    "snippet_source": p["content"],
                    "thread_size": 1,
                })
        return matched


# ==================== 計算邏輯 ====================

def categorize_station(stb):
    """把站版名稱對應到模板分類"""
    s = normalize(stb)
    if "PTT" in s.upper():
        return "PTT"
    if "Dcard" in s or "狄卡" in s:
        return "Dcard"
    if "Threads" in s:
        return "Threads"
    if "Facebook" in s or "FB" in s:
        return "FB社團"
    return "其他版面"


def calculate_all(main_data, post_types):
    """計算所有要填入模板的數字"""
    result = {}
    
    # ===== 工作表 1:篇數 =====
    type_mapping = {
        "2~3圖+分享文(素人拍+過稿)": ["2~3圖+分享文(提供照片+過稿)", "2~3圖+分享文(素人拍+過稿)", "2~3圖+分享文"],
        "1圖+分享文(素人拍+過稿)": ["1圖+分享文(提供照片+過稿)", "1圖+分享文(素人拍+過稿)", "1圖+分享文"],
        "分享文": ["分享文"],
        "詳文": ["詳文"],
        "主文": ["主文"],
        "推文": ["推文"],
        "FB-主文": ["FB-主文"],
        "FB-推文": ["FB-推文"],
        "置入主文": ["置入主文"],
        "置入推文": ["置入推文"],
    }
    
    篇數 = {}
    for template_name, pdf_names in type_mapping.items():
        value = 0
        for pdf_name in pdf_names:
            for k, v in post_types.items():
                if normalize(pdf_name) == normalize(k):
                    value = v
                    break
            if value > 0:
                break
        篇數[template_name] = value
    result["篇數"] = 篇數
    
    # ===== 工作表 2:KPI =====
    kpi = {}
    kpi["網友回應數"] = sum(r["網友回應量"] for r in main_data)
    kpi["議題曝光數"] = sum(1 for r in main_data if "置入" not in r["主軸"])
    kpi["內文指名度"] = None
    kpi["好評增加數"] = sum(r["正向聲量總數"] - r["正面討論"] for r in main_data)
    result["KPI"] = kpi
    
    # ===== 工作表 3:網友回應分布 =====
    分布 = {}
    for cat in ["PTT", "Dcard", "Threads", "其他版面", "FB社團"]:
        items = [r for r in main_data if categorize_station(r["站版"]) == cat]
        分布[cat] = {
            "實際溝通": sum(r["專案發文量"] for r in items),
            "正面討論": sum(r["正面討論"] for r in items),
            "負面討論": sum(r["負面討論"] for r in items),
            "議題討論": sum(r["議題討論"] for r in items),
            "產品討論": sum(r["產品討論"] for r in items),
            "複合討論": sum(r["複合討論"] for r in items),
        }
    result["分布"] = 分布
    result["分布來源"] = {
        cat: [r["站版"] for r in main_data if categorize_station(r["站版"]) == cat]
        for cat in ["PTT", "Dcard", "Threads", "其他版面", "FB社團"]
    }
    
    return result


# ==================== 月報表填充邏輯 ====================

def categorize_station_monthly(stb):
    """月報表的站版分類"""
    s = normalize(stb).lower()
    if "ptt" in s or "批踢踢" in stb:
        return "PTT"
    if "dcard" in s or "狄卡" in stb:
        return "DCARD"
    if "thread" in s:
        return "THREAD"
    return "其他版面"


def fill_monthly_template(template_path, main_data, post_types):
    """填入月報表模板"""
    wb = load_workbook(template_path)
    
    total_post = sum(r["專案發文量"] for r in main_data)
    total_reply = sum(r["網友回應量"] for r in main_data)
    total_volume = sum(r["討論聲量總數"] for r in main_data)
    sum_pos = sum(r["正面討論"] for r in main_data)
    sum_neg = sum(r["負面討論"] for r in main_data)
    sum_topic = sum(r["議題討論"] for r in main_data)
    sum_prod = sum(r["產品討論"] for r in main_data)
    sum_combo = sum(r["複合討論"] for r in main_data)
    
    denominator = sum_pos + sum_neg + sum_topic + sum_prod + sum_combo
    
    def pct(numerator):
        if denominator == 0:
            return 0.0
        return round(numerator / denominator * 100, 1)
        
    ws1 = wb["網友回應摘要+操作數據"]
    
    type_detail_parts = []
    for tname, count in post_types.items():
        type_detail_parts.append(f"{tname}{count}")
    if type_detail_parts:
        type_detail_str = "/".join(type_detail_parts)
        ws1["C6"] = f"{total_post}篇\n【{type_detail_str}】"
    else:
        ws1["C6"] = f"{total_post}篇"
        
    ws1["C7"] = f"{total_reply}篇"
    ws1["C8"] = f"{total_volume}篇"
    ws1["C9"] = f"{pct(sum_pos)}% / {sum_pos}篇"
    ws1["C10"] = f"{pct(sum_neg)}% / {sum_neg}篇"
    ws1["C11"] = f"{pct(sum_combo)}% / {sum_combo}篇"
    ws1["C12"] = f"{pct(sum_prod)}% / {sum_prod}篇"
    ws1["C13"] = f"{pct(sum_topic)}% / {sum_topic}篇"
    
    ws2 = wb["版面佔比"]
    cat_reply_sum = {"PTT": 0, "DCARD": 0, "THREAD": 0, "其他版面": 0}
    for r in main_data:
        cat = categorize_station_monthly(r["站版"])
        cat_reply_sum[cat] += r["網友回應量"]
        
    ws2["C10"] = cat_reply_sum["PTT"]
    ws2["C11"] = cat_reply_sum["DCARD"]
    ws2["C12"] = cat_reply_sum["THREAD"]
    ws2["C13"] = cat_reply_sum["其他版面"]
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def fill_template(template_path, calc_result):
    """把計算結果填入模板"""
    wb = load_workbook(template_path)
    
    ws1 = wb["篇數"]
    篇數欄位映射 = {
        "2~3圖+分享文(素人拍+過稿)": "C8",
        "1圖+分享文(素人拍+過稿)": "D8",
        "分享文": "E8",
        "詳文": "F8",
        "主文": "G8",
        "推文": "H8",
        "FB-主文": "I8",
        "FB-推文": "J8",
        "置入主文": "K8",
        "置入推文": "L8",
    }
    for name, cell in 篇數欄位映射.items():
        ws1[cell] = calc_result["篇數"].get(name, 0)
        
    上方小表映射 = {
        "2~3圖+分享文(素人拍+過稿)": "C4",
        "1圖+分享文(素人拍+過稿)": "D4",
        "分享文": "E4",
        "詳文": "F4",
        "主文": "G4",
        "推文": "H4",
        "FB-主文": "I4",
        "FB-推文": "J4",
        "置入主文": "K4",
        "置入推文": "L4",
    }
    for name, cell in 上方小表映射.items():
        ws1[cell] = calc_result["篇數"].get(name, 0)
    ws1["C5"] = "=SUM(C4:L4)"
    
    區塊B映射 = {
        "2~3圖+分享文(素人拍+過稿)": "C15",
        "1圖+分享文(素人拍+過稿)": "C16",
        "分享文": "C17",
        "詳文": "C18",
        "主文": "C19",
        "推文": "C20",
        "FB-主文": "C21",
        "FB-推文": "C22",
        "置入主文": "C23",
        "置入推文": "C24",
    }
    for name, cell in 區塊B映射.items():
        ws1[cell] = calc_result["篇數"].get(name, 0)
    ws1["C25"] = "=SUM(C15:C24)"
    
    ws2 = wb["KPI"]
    ws2["D4"] = f"{calc_result['KPI']['網友回應數']}篇"
    ws2["D5"] = f"{calc_result['KPI']['議題曝光數']}串"
    ws2["D6"] = ""
    ws2["D7"] = f"{calc_result['KPI']['好評增加數']}篇"
    
    ws3 = wb["網友回應分布"]
    版面列對應 = {
        "PTT": 5,
        "Dcard": 6,
        "Threads": 7,
        "其他版面": 8,
        "FB社團": 9,
    }
    for cat, row_num in 版面列對應.items():
        d = calc_result["分布"][cat]
        ws3.cell(row=row_num, column=3).value = d["實際溝通"]
        ws3.cell(row=row_num, column=4).value = d["正面討論"]
        ws3.cell(row=row_num, column=5).value = d["負面討論"]
        ws3.cell(row=row_num, column=6).value = d["議題討論"]
        ws3.cell(row=row_num, column=7).value = d["產品討論"]
        ws3.cell(row=row_num, column=8).value = d["複合討論"]
        ws3.cell(row=row_num, column=9).value = f"=SUM(D{row_num}:H{row_num})"
        
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ==================== 預覽渲染輔助函式 ====================

def _render_closure_preview(result, main_data, post_types):
    """渲染結案表預覽"""
    tab1, tab2, tab3, tab4 = st.tabs(["📑 篇數", "📈 KPI", "📊 網友回應分布", "📂 原始資料"])
    
    with tab1:
        st.write("**將填入工作表「篇數」**")
        篇數 = result["篇數"]
        cols = st.columns(5)
        for i, (k, v) in enumerate(篇數.items()):
            cols[i % 5].metric(k, v)
        st.caption(f"📍 PDF 中抓到的文案類型:{', '.join(post_types.keys()) if post_types else '(無)'}")
        
    with tab2:
        st.write("**將填入工作表「KPI」**")
        kpi = result["KPI"]
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("網友回應數", f"{kpi['網友回應數']} 篇")
        c2.metric("議題曝光數", f"{kpi['議題曝光數']} 串")
        c3.metric("內文指名度", "(手動填)", delta="留空")
        c4.metric("好評增加數", f"{kpi['好評增加數']} 篇")
        
        with st.expander("📐 計算明細"):
            st.write(f"- **網友回應數** = 網友回應量加總 = **{kpi['網友回應數']}**")
            non_p = [r['主軸'] for r in main_data if "置入" not in r['主軸']]
            st.write(f"- **議題曝光數** = 主軸非「置入」的列數 = **{kpi['議題曝光數']}**")
            st.write(f"  非置入主軸明細: {non_p}")
            row_diffs = [(r['站版'][:20], r['正向聲量總數'], r['正面討論'], r['正向聲量總數']-r['正面討論']) for r in main_data]
            st.write(f"- **好評增加數** = 各列(正向聲量−正面討論)先計算再加總 = **{kpi['好評增加數']}**")
            with st.expander("查看每列計算"):
                import pandas as pd
                df_diff = pd.DataFrame(row_diffs, columns=["站版", "正向聲量", "正面討論", "差值"])
                st.dataframe(df_diff, use_container_width=True, hide_index=True)
                
    with tab3:
        st.write("**將填入工作表「網友回應分布」**")
        import pandas as pd
        df_dist = pd.DataFrame(result["分布"]).T
        df_dist["回應小計"] = df_dist[["正面討論", "負面討論", "議題討論", "產品討論", "複合討論"]].sum(axis=1)
        df_dist = df_dist[["實際溝通", "正面討論", "負面討論", "議題討論", "產品討論", "複合討論", "回應小計"]]
        st.dataframe(df_dist, use_container_width=True)
        
        with st.expander("🗂️ 各分類包含的站版"):
            for cat, sources in result["分布來源"].items():
                if sources:
                    st.write(f"**{cat}**: {', '.join(set(sources))}")
                else:
                    st.write(f"**{cat}**: (無)")
                    
    with tab4:
        st.write("**從 PDF 抽出的原始明細**")
        import pandas as pd
        df_raw = pd.DataFrame(main_data)
        st.dataframe(df_raw, use_container_width=True, hide_index=True)


def _render_monthly_preview(main_data, post_types):
    """渲染月報表預覽"""
    tab1, tab2, tab3 = st.tabs(["📑 操作數據", "📊 版面佔比", "📂 原始資料"])
    
    total_post = sum(r["專案發文量"] for r in main_data)
    total_reply = sum(r["網友回應量"] for r in main_data)
    total_volume = sum(r["討論聲量總數"] for r in main_data)
    sum_pos = sum(r["正面討論"] for r in main_data)
    sum_neg = sum(r["負面討論"] for r in main_data)
    sum_topic = sum(r["議題討論"] for r in main_data)
    sum_prod = sum(r["產品討論"] for r in main_data)
    sum_combo = sum(r["複合討論"] for r in main_data)
    denominator = sum_pos + sum_neg + sum_topic + sum_prod + sum_combo
    
    def pct(n):
        return round(n / denominator * 100, 1) if denominator else 0.0
        
    with tab1:
        st.write("**將填入頁簽「網友回應摘要+操作數據」**")
        c1, c2, c3 = st.columns(3)
        c1.metric("專案總發文量", f"{total_post} 篇")
        c2.metric("網友總回應量", f"{total_reply} 篇")
        c3.metric("討論聲量總數", f"{total_volume} 篇")
        
        st.write("**各討論類型比率/篇數**")
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("正面", f"{pct(sum_pos)}%", f"{sum_pos} 篇")
        c2.metric("負面", f"{pct(sum_neg)}%", f"{sum_neg} 篇")
        c3.metric("複合", f"{pct(sum_combo)}%", f"{sum_combo} 篇")
        c4.metric("產品", f"{pct(sum_prod)}%", f"{sum_prod} 篇")
        c5.metric("議題", f"{pct(sum_topic)}%", f"{sum_topic} 篇")
        
        if post_types:
            with st.expander("📑 文案類型明細"):
                import pandas as pd
                df_pt = pd.DataFrame([(k, v) for k, v in post_types.items()], columns=["文案類型", "篇數"])
                st.dataframe(df_pt, use_container_width=True, hide_index=True)
                
    with tab2:
        st.write("**將填入頁簽「版面佔比」**(網友回應量加總)")
        cat_reply_sum = {"PTT": 0, "DCARD": 0, "THREAD": 0, "其他版面": 0}
        cat_sources = {"PTT": [], "DCARD": [], "THREAD": [], "其他版面": []}
        for r in main_data:
            cat = categorize_station_monthly(r["站版"])
            cat_reply_sum[cat] += r["網友回應量"]
            cat_sources[cat].append(r["站版"])
            
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("PTT", cat_reply_sum["PTT"])
        c2.metric("DCARD", cat_reply_sum["DCARD"])
        c3.metric("THREAD", cat_reply_sum["THREAD"])
        c4.metric("其他版面", cat_reply_sum["其他版面"])
        
        with st.expander("🗂️ 各分類包含的站版"):
            for cat, sources in cat_sources.items():
                unique_sources = set(sources)
                if unique_sources:
                    st.write(f"**{cat}**: {', '.join(unique_sources)}")
                else:
                    st.write(f"**{cat}**: (無)")
                    
        with st.expander("🔍 除錯:每一筆的分類明細"):
            import pandas as pd
            debug_rows = []
            for i, r in enumerate(main_data, 1):
                cat = categorize_station_monthly(r["站版"])
                debug_rows.append({
                    "#": i,
                    "站版字串(原始)": repr(r["站版"])[:60],
                    "分類結果": cat,
                    "網友回應量": r["網友回應量"],
                })
            df_debug = pd.DataFrame(debug_rows)
            st.dataframe(df_debug, use_container_width=True, hide_index=True)
            
    with tab3:
        st.write("**從 PDF 抽出的原始明細**")
        import pandas as pd
        df_raw = pd.DataFrame(main_data)
        st.dataframe(df_raw, use_container_width=True, hide_index=True)


# ==================== Streamlit UI ====================

st.subheader("產出報表設定")
col_type, col_kw, col_scope = st.columns([1.5, 2.5, 1])

with col_type:
    report_type = st.radio(
        "選擇功能",
        ["結案表", "月報表", "內文指名度查詢"],
        horizontal=False,
        key="report_type_ui"
    )

with col_kw:
    keyword = st.text_input(
        "關鍵字搜尋 (限指名度查詢功能)",
        placeholder="例如: Friday、Friday+稻草人",
        key="search_keyword",
        disabled=(report_type != "內文指名度查詢")
    )

with col_scope:
    search_in = st.radio(
        "搜尋範圍",
        ["內文", "標題"],
        key="search_scope",
        disabled=(report_type != "內文指名度查詢")
    )

search_in_key = "content" if search_in == "內文" else "title"

st.divider()

# === Step 2:上傳 PDF ===
st.subheader("📤 上傳 PDF 報表")
pdf_file = st.file_uploader(
    "銀河專案 PDF",
    type=["pdf"],
    key="pdf",
)

if pdf_file:
    st.divider()
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
        tmp_pdf.write(pdf_file.getvalue())
        pdf_path = tmp_pdf.name
        
    try:
        # ===== 分支:內文指名度查詢 =====
        if report_type == "內文指名度查詢":
            with st.spinner("🔍 解析 PDF 文章..."):
                posts = extract_posts_with_content(pdf_path)
                
            if not posts:
                st.error("❌ 無法從 PDF 抽出文章。請確認上傳的是『專案發文總覽』PDF")
                st.stop()
                
            st.success(f"✅ 解析成功!共抓到 **{len(posts)}** 篇文章")
            
            total_content_chars = sum(len(p["content"]) for p in posts)
            total_title_chars = sum(len(p["title"]) for p in posts)
            c1, c2, c3 = st.columns(3)
            c1.metric("文章篇數", len(posts))
            c2.metric("內文總字數", f"{total_content_chars:,}")
            c3.metric("標題總字數", f"{total_title_chars:,}")
            
            if keyword.strip():
                matched = search_posts(posts, keyword, search_in=search_in_key)
                
                if search_in_key == "title":
                    from collections import OrderedDict
                    thread_set = OrderedDict()
                    for p in posts:
                        key = normalize_thread_title(p["title"])
                        thread_set[key] = True
                    total_base = len(thread_set)
                    base_label = "總討論串數"
                else:
                    total_base = len(posts)
                    base_label = "總篇數"
                    
                st.divider()
                st.markdown(f"### 📊 搜尋結果")
                col_r1, col_r2, col_r3 = st.columns(3)
                unit = "串" if search_in_key == "title" else "篇"
                col_r1.metric(f"在「{search_in}」中命中", f"{len(matched)} {unit}")
                col_r2.metric(base_label, f"{total_base} {unit}")
                pct_val = (len(matched) / total_base * 100) if total_base else 0
                col_r3.metric("命中比例", f"{pct_val:.1f}%")
                
                if matched:
                    label_word = "討論串" if search_in_key == "title" else "文章"
                    st.subheader(f"📋 命中{label_word}清單")
                    import pandas as pd
                    rows = []
                    for i, p in enumerate(matched, 1):
                        target = p["snippet_source"]
                        first_kw = re.split(r'[+,|]', keyword.strip())[0].strip()
                        idx = target.lower().find(first_kw.lower())
                        if idx >= 0:
                            ctx_start = max(0, idx - 20)
                            ctx_end = min(len(target), idx + len(first_kw) + 40)
                            snippet = ("..." if ctx_start > 0 else "") + target[ctx_start:ctx_end].replace('\n', ' ') + ("..." if ctx_end < len(target) else "")
                        else:
                            snippet = target[:60].replace('\n', ' ') + ("..." if len(target) > 60 else "")
                            
                        row = {
                            "#": i,
                            "標題": p["title"][:60],
                            f"{search_in}片段": snippet,
                        }
                        if search_in_key == "title" and p.get("thread_size", 1) > 1:
                            row["該串包含文章數"] = p["thread_size"]
                        rows.append(row)
                    df_results = pd.DataFrame(rows)
                    st.dataframe(df_results, use_container_width=True, hide_index=True)
                    
                    csv_data = df_results.to_csv(index=False).encode('utf-8-sig')
                    st.download_button(
                        "📥 下載命中清單(CSV)",
                        data=csv_data,
                        file_name=f"內文指名度_{keyword.replace('+','_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                    )
                else:
                    st.info("沒有任何文章轉換符合此關鍵字條件")
            else:
                st.info("👆 請於上方輸入關鍵字以執行指名度查詢")
                
        # ===== 分支:報表填充功能 =====
        else:
            try:
                template_bytes = get_builtin_template(report_type)
            except Exception as e:
                st.error(f"❌ 載入內建模板失敗: {e}")
                st.stop()
                
            with st.spinner("🔍 解析 PDF 並計算數據..."):
                main_data = extract_main_table(pdf_path)
                post_types = extract_post_types(pdf_path)
                
            if not main_data:
                st.error("❌ 無法從 PDF 抽取主表資料，請確認上傳格式。")
                st.stop()
                
            st.success(f"✅ 解析成功!共抓到 **{len(main_data)}** 筆資料")
            
            st.subheader("👀 預覽計算結果")
            if report_type == "結案表":
                result = calculate_all(main_data, post_types)
                _render_closure_preview(result, main_data, post_types)
            else:
                _render_monthly_preview(main_data, post_types)
                
            st.divider()
            
            with st.spinner("📝 填入模板..."):
                try:
                    if report_type == "月報表":
                        filled_bytes = fill_monthly_template(BytesIO(template_bytes), main_data, post_types)
                    else:
                        filled_bytes = fill_template(BytesIO(template_bytes), result)
                except Exception as e:
                    st.error(f"❌ 填入模板失敗: {e}")
                    st.stop()
                    
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_name = f"{report_type}_已填_{timestamp}.xlsx"
            
            st.download_button(
                label=f"📥 下載填好的 {report_type}",
                data=filled_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
    finally:
        if os.path.exists(pdf_path):
            os.unlink(pdf_path)
else:
    st.info("👆 請選擇功能並上傳對應 PDF 報告檔案以開始執行。")
